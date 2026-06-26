import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelFormatter:
    def __init__(self):
        # Setup warna pastel agar terlihat profesional (tidak merusak mata)
        self.fill_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Akumulasi
        self.fill_red = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')   # Distribusi
        self.fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Biru Navy
        
        self.font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        self.font_data = Font(name='Segoe UI', size=10)
        self.font_bold = Font(name='Segoe UI', size=10, bold=True)

    def write_report(self, df_analysis: pd.DataFrame, output_path: str):
        """
        Menulis DataFrame hasil analisis ke file Excel dengan format otomatis yang cantik.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Whale Tracker Harian"
        ws.views.sheetView[0].showGridLines = True

        # 1. Tulis Header Custom
        headers = [
            'Ticker', 'Nama Emiten', 'Pemegang Saham', 
            'Saham Kemarin', 'Saham Hari Ini', 'Perubahan Saham', 
            '% Kemarin', '% Hari Ini', 'Delta %', 'Status Analisis'
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = self.fill_header
            cell.font = self.font_header
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 2. Tulis Data & Terapkan Conditional Formatting secara Row-by-Row
        for row_idx, row_data in enumerate(dataframe_to_rows(df_analysis, index=False, header=False), start=2):
            ws.append(row_data)
            
            status = row_data[9] # Kolom Status Analisis ada di indeks ke-9
            current_row = ws[row_idx]
            
            # Tentukan warna berdasarkan status akumulasi / distribusi
            row_fill = None
            if status in ['AKUMULASI', 'NEW_ENTRY']:
                row_fill = self.fill_green
            elif status in ['DISTRIBUSI', 'EXIT']:
                row_fill = self.fill_red
                
            # Formatting tiap cell di baris ini
            for col_idx, cell in enumerate(current_row, start=1):
                cell.font = self.font_data
                if row_fill:
                    cell.fill = row_fill
                
                # Format angka ribuan untuk jumlah saham (Kolom D, E, F)
                if col_idx in [4, 5, 6]:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                # Format persentase (Kolom G, H, I)
                elif col_idx in [7, 8, 9]:
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal='right')
                elif col_idx in [1, 10]:
                    cell.alignment = Alignment(horizontal='center')

        # 3. Auto-fit lebar kolom biar tidak kepotong (###)
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Save file Excel
        wb.save(output_path)