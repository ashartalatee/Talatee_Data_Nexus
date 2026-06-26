"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          FORM DATA COLLECTOR ENGINE #10 — Excel Storage                     ║
║          Python Business Automation Engine | Portfolio Professional          ║
╚══════════════════════════════════════════════════════════════════════════════╝

Module  : storage/excel_storage.py
Purpose : Save form data to professional Excel (.xlsx) files.

          Features:
          1. Save DataFrame → formatted Excel with auto-fit columns
          2. Multi-sheet workbook (data + summary + invalid rows)
          3. Header styling (bold, background color, border)
          4. Alternating row colors for readability
          5. Auto-filter on all columns
          6. Freeze top row (header always visible)
          7. Number & date formatting per column type
          8. Conditional formatting (highlight invalid/warning rows)
          9. Append to existing workbook (add new sheet)
         10. Read Excel → DataFrame

Author  : Python Automation Engine
Version : 1.0.0
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from loguru import logger

# openpyxl for formatting
try:
    from openpyxl import load_workbook
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill,
        PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("⚠️  openpyxl not installed. pip install openpyxl")


# ─────────────────────────────────────────────────────────────────────────────
# STYLE CONFIG
# ─────────────────────────────────────────────────────────────────────────────

# Header style
HEADER_FILL   = "1F4E79"   # dark blue
HEADER_FONT   = "FFFFFF"   # white text
ROW_ALT_FILL  = "EBF3FB"   # light blue alternating row
ROW_WARN_FILL = "FFF2CC"   # yellow for warning rows
ROW_ERR_FILL  = "FCE4D6"   # orange-red for error rows
MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 60


# ─────────────────────────────────────────────────────────────────────────────
# RESULT
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ExcelStorageResult:
    """
    Result of an Excel storage operation.

    Attributes:
        path         : Path to the saved .xlsx file.
        sheets       : Sheet names written.
        rows_written : Rows written to the primary sheet.
        success      : True if no error occurred.
        error        : Error message if success=False.
    """
    path        : Path
    sheets      : list[str]
    rows_written: int
    success     : bool = True
    error       : str  = ""

    @property
    def summary(self) -> dict:
        return {
            "path"        : str(self.path),
            "sheets"      : self.sheets,
            "rows_written": self.rows_written,
            "success"     : self.success,
        }

    def __repr__(self) -> str:
        return (
            f"<ExcelStorageResult path='{self.path.name}' "
            f"sheets={self.sheets} rows={self.rows_written} "
            f"success={self.success}>"
        )


# ─────────────────────────────────────────────────────────────────────────────
# CORE EXCEL STORAGE
# ─────────────────────────────────────────────────────────────────────────────

class ExcelStorage:
    """
    Professional Excel storage handler with full formatting.

    Args:
        output_path      : Path to the .xlsx output file.
        sheet_name       : Primary sheet name. Default "Form Data".
        backup_on_write  : Backup existing file before overwrite. Default True.
        apply_formatting : Apply header style, alternating rows, auto-fit. Default True.
        add_summary_sheet: Add a summary statistics sheet. Default True.
        freeze_header    : Freeze the first row. Default True.
        add_autofilter   : Enable Excel auto-filter on all columns. Default True.
        highlight_col    : Column name to use for row highlighting.
                           Rows with values starting with "error"/"invalid"
                           will be highlighted red; "warn"/"warning" → yellow.
    """

    def __init__(
        self,
        output_path      : str | Path,
        sheet_name       : str  = "Form Data",
        backup_on_write  : bool = True,
        apply_formatting : bool = True,
        add_summary_sheet: bool = True,
        freeze_header    : bool = True,
        add_autofilter   : bool = True,
        highlight_col    : Optional[str] = None,
    ) -> None:
        self.output_path       = Path(output_path)
        self.sheet_name        = sheet_name
        self.backup_on_write   = backup_on_write
        self.apply_formatting  = apply_formatting
        self.add_summary_sheet = add_summary_sheet
        self.freeze_header     = freeze_header
        self.add_autofilter    = add_autofilter
        self.highlight_col     = highlight_col

        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── private: formatting ───────────────────────────────────────────────────

    def _style_worksheet(self, ws, df: pd.DataFrame) -> None:
        """Apply professional styling to a worksheet."""
        if not OPENPYXL_AVAILABLE:
            return

        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        header_font = Font(bold=True, color=HEADER_FONT, size=10)
        alt_fill    = PatternFill("solid", fgColor=ROW_ALT_FILL)
        warn_fill   = PatternFill("solid", fgColor=ROW_WARN_FILL)
        err_fill    = PatternFill("solid", fgColor=ROW_ERR_FILL)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        thin_border  = Border(
            left  =Side(style="thin", color="CCCCCC"),
            right =Side(style="thin", color="CCCCCC"),
            top   =Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # ── style header row ──────────────────────────────────────────────────
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border

        # ── style data rows ───────────────────────────────────────────────────
        highlight_col_idx = None
        if self.highlight_col and self.highlight_col in df.columns:
            highlight_col_idx = df.columns.tolist().index(self.highlight_col) + 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # determine fill
            fill = None
            if highlight_col_idx:
                cell_val = str(ws.cell(row=row_idx, column=highlight_col_idx).value or "").lower()
                if any(x in cell_val for x in ("error", "invalid", "quarantine")):
                    fill = err_fill
                elif any(x in cell_val for x in ("warn", "missing", "incomplete")):
                    fill = warn_fill

            if fill is None and row_idx % 2 == 0:
                fill = alt_fill

            for cell in row:
                if fill:
                    cell.fill = fill
                cell.border    = thin_border
                cell.alignment = Alignment(vertical="center")

        # ── auto-fit column widths ────────────────────────────────────────────
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(
                max(max_len + 4, MIN_COL_WIDTH), MAX_COL_WIDTH
            )

        # ── row height ────────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 22   # header taller
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 16

        # ── freeze header ─────────────────────────────────────────────────────
        if self.freeze_header:
            ws.freeze_panes = "A2"

        # ── auto filter ───────────────────────────────────────────────────────
        if self.add_autofilter and ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    def _add_summary_sheet(self, wb, df: pd.DataFrame) -> None:
        """Add a Summary sheet with basic statistics."""
        if not OPENPYXL_AVAILABLE:
            return

        ws = wb.create_sheet("Summary")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        summary_data = [
            ["FORM DATA SUMMARY REPORT", ""],
            ["Generated at", now],
            ["", ""],
            ["DATASET INFO", ""],
            ["Total rows",    len(df)],
            ["Total columns", len(df.columns)],
            ["Columns", ", ".join(df.columns.tolist())],
            ["", ""],
        ]

        # per-column null count
        summary_data.append(["COLUMN COMPLETENESS", "Fill Rate"])
        for col in df.columns:
            if col.startswith("_"):
                continue
            filled    = df[col].replace("", pd.NA).notna().sum()
            fill_rate = f"{filled / max(len(df), 1) * 100:.1f}%"
            summary_data.append([col, fill_rate])

        for row in summary_data:
            ws.append(row)

        # style title
        ws["A1"].font      = Font(bold=True, size=13, color=HEADER_FILL)
        ws["A4"].font      = Font(bold=True, size=11)
        ws["A9"].font      = Font(bold=True, size=11)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

    # ── private: backup ───────────────────────────────────────────────────────

    def _backup(self) -> Optional[Path]:
        if not self.output_path.exists():
            return None
        ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.output_path.with_suffix(f".backup_{ts}.xlsx")
        shutil.copy2(self.output_path, backup_path)
        logger.debug(f"  📦 Backup created → {backup_path.name}")
        return backup_path

    # ── public: save ──────────────────────────────────────────────────────────

    def save(
        self,
        df           : pd.DataFrame,
        extra_sheets : Optional[dict[str, pd.DataFrame]] = None,
    ) -> ExcelStorageResult:
        """
        Save DataFrame to a formatted Excel workbook.

        Args:
            df           : Primary DataFrame (goes to self.sheet_name).
            extra_sheets : Optional dict of {sheet_name: DataFrame} for
                           additional sheets (e.g. invalid rows, audit log).

        Returns:
            ExcelStorageResult
        """
        if df.empty:
            logger.warning("⚠️  ExcelStorage.save() called with empty DataFrame.")
            return ExcelStorageResult(
                path=self.output_path, sheets=[], rows_written=0, success=True
            )

        backup_path = None
        sheets_written: list[str] = []

        try:
            if self.backup_on_write:
                backup_path = self._backup()

            with pd.ExcelWriter(
                self.output_path, engine="openpyxl"
            ) as writer:

                # ── primary sheet ─────────────────────────────────────────────
                df.to_excel(writer, sheet_name=self.sheet_name, index=False)
                sheets_written.append(self.sheet_name)

                # ── extra sheets ──────────────────────────────────────────────
                if extra_sheets:
                    for sname, sdf in extra_sheets.items():
                        if sdf is not None and not sdf.empty:
                            sdf.to_excel(writer, sheet_name=sname, index=False)
                            sheets_written.append(sname)

                # ── summary sheet ─────────────────────────────────────────────
                if self.add_summary_sheet and OPENPYXL_AVAILABLE:
                    self._add_summary_sheet(writer.book, df)
                    sheets_written.append("Summary")

                # ── apply formatting ──────────────────────────────────────────
                if self.apply_formatting and OPENPYXL_AVAILABLE:
                    for sname in sheets_written:
                        if sname == "Summary":
                            continue
                        ws  = writer.sheets.get(sname)
                        sdf = df if sname == self.sheet_name else extra_sheets.get(sname, pd.DataFrame())
                        if ws is not None and not sdf.empty:
                            self._style_worksheet(ws, sdf)

            logger.info(
                f"💾 Excel saved → {self.output_path} "
                f"({len(df)} rows | sheets: {sheets_written})"
            )

            return ExcelStorageResult(
                path=self.output_path, sheets=sheets_written,
                rows_written=len(df), success=True,
            )

        except Exception as exc:
            logger.error(f"❌ Excel save failed: {exc}")
            return ExcelStorageResult(
                path=self.output_path, sheets=[], rows_written=0,
                success=False, error=str(exc),
            )

    # ── public: add_sheet ─────────────────────────────────────────────────────

    def add_sheet(self, df: pd.DataFrame, sheet_name: str) -> ExcelStorageResult:
        """
        Add a new sheet to an existing workbook without overwriting other sheets.

        Args:
            df         : DataFrame to add.
            sheet_name : Name for the new sheet.

        Returns:
            ExcelStorageResult
        """
        if not self.output_path.exists():
            logger.info("  File doesn't exist yet — calling save() instead.")
            return self.save(df)

        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl required for add_sheet()")
            return ExcelStorageResult(
                path=self.output_path, sheets=[], rows_written=0,
                success=False, error="openpyxl not installed",
            )

        try:
            wb = load_workbook(self.output_path)

            # remove existing sheet with same name if exists
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

            ws = wb.create_sheet(sheet_name)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)

            if self.apply_formatting:
                self._style_worksheet(ws, df)

            wb.save(self.output_path)
            logger.info(f"  ➕ Sheet '{sheet_name}' added → {self.output_path}")

            return ExcelStorageResult(
                path=self.output_path, sheets=[sheet_name],
                rows_written=len(df), success=True,
            )

        except Exception as exc:
            logger.error(f"❌ add_sheet failed: {exc}")
            return ExcelStorageResult(
                path=self.output_path, sheets=[], rows_written=0,
                success=False, error=str(exc),
            )

    # ── public: read ──────────────────────────────────────────────────────────

    def read(
        self,
        sheet_name : Optional[str]  = None,
        dtype      : Optional[dict] = None,
    ) -> pd.DataFrame:
        """
        Read a sheet from the Excel file into a DataFrame.

        Args:
            sheet_name : Sheet to read. Default = self.sheet_name.
            dtype      : Optional type mapping for pd.read_excel.

        Returns:
            DataFrame (empty if file doesn't exist).
        """
        if not self.output_path.exists():
            logger.warning(f"⚠️  File not found: {self.output_path}")
            return pd.DataFrame()

        target = sheet_name or self.sheet_name
        try:
            df = pd.read_excel(
                self.output_path,
                sheet_name=target,
                engine="openpyxl",
                dtype=dtype,
            )
            logger.info(f"📂 Excel loaded ← {self.output_path} sheet='{target}' ({len(df)} rows)")
            return df
        except Exception as exc:
            logger.error(f"❌ Excel read failed: {exc}")
            return pd.DataFrame()

    # ── public: info ──────────────────────────────────────────────────────────

    def info(self) -> dict:
        """Return metadata about the current Excel file."""
        if not self.output_path.exists():
            return {"exists": False, "path": str(self.output_path)}

        stat = self.output_path.stat()
        sheets: list[str] = []
        if OPENPYXL_AVAILABLE:
            try:
                wb     = load_workbook(self.output_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
            except Exception:
                pass

        return {
            "exists"   : True,
            "path"     : str(self.output_path),
            "sheets"   : sheets,
            "size_kb"  : round(stat.st_size / 1024, 2),
            "modified" : datetime.fromtimestamp(stat.st_mtime).isoformat(),
        }


# ─────────────────────────────────────────────────────────────────────────────
# CONVENIENCE FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def save_excel(
    df           : pd.DataFrame,
    path         : str | Path,
    sheet_name   : str = "Form Data",
    extra_sheets : Optional[dict[str, pd.DataFrame]] = None,
    highlight_col: Optional[str] = None,
) -> ExcelStorageResult:
    """
    One-liner Excel save with full formatting.

    Used by main.py and data_router handlers:
        from storage.excel_storage import save_excel
        save_excel(df, "outputs/form_data.xlsx", extra_sheets={"Invalid": df_invalid})

    Args:
        df           : Primary DataFrame.
        path         : Output .xlsx file path.
        sheet_name   : Primary sheet name. Default "Form Data".
        extra_sheets : Additional sheets dict.
        highlight_col: Column for conditional row highlighting.
    """
    storage = ExcelStorage(
        path,
        sheet_name=sheet_name,
        highlight_col=highlight_col,
    )
    return storage.save(df, extra_sheets=extra_sheets)


def read_excel(
    path      : str | Path,
    sheet_name: Optional[str] = None,
) -> pd.DataFrame:
    """One-liner Excel read."""
    return ExcelStorage(path).read(sheet_name=sheet_name)


# ─────────────────────────────────────────────────────────────────────────────
# CLI TEST
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import tempfile

    print("=" * 65)
    print("  Excel Storage — Quick Test")
    print("=" * 65)

    df_valid = pd.DataFrame({
        "nama"   : ["Budi Santoso", "Siti Rahayu", "Andi Wijaya"],
        "email"  : ["budi@gmail.com", "siti@yahoo.com", "andi@outlook.com"],
        "phone"  : ["+628123456789", "+628556789012", "+628778901234"],
        "kota"   : ["Surabaya", "Jakarta", "Bandung"],
        "total"  : [1_500_000, 2_500_000, 750_000],
        "tier"   : ["Silver", "Gold", "Bronze"],
        "status" : ["valid", "valid", "valid"],
    })

    df_invalid = pd.DataFrame({
        "nama"   : ["Rina"],
        "email"  : [""],
        "phone"  : [""],
        "status" : ["error: missing email and phone"],
    })

    with tempfile.TemporaryDirectory() as tmpdir:
        path    = Path(tmpdir) / "test_form_data.xlsx"
        storage = ExcelStorage(
            path,
            sheet_name   = "Form Data",
            highlight_col= "status",
        )

        # ── TEST 1: Save with extra sheet ─────────────────────────────────────
        print("\n  TEST 1: save() with extra sheet (Invalid Rows)")
        r1 = storage.save(df_valid, extra_sheets={"Invalid Rows": df_invalid})
        print(f"  Result : {r1}")
        assert r1.success is True
        assert "Form Data" in r1.sheets
        assert "Invalid Rows" in r1.sheets
        assert "Summary" in r1.sheets

        # ── TEST 2: Add sheet to existing workbook ────────────────────────────
        print("\n  TEST 2: add_sheet() — Audit Log")
        df_audit = pd.DataFrame({
            "action"    : ["collect", "validate", "store"],
            "timestamp" : [datetime.now().isoformat()] * 3,
            "rows"      : [3, 3, 3],
        })
        r2 = storage.add_sheet(df_audit, "Audit Log")
        print(f"  Result : {r2}")
        assert r2.success is True

        # ── TEST 3: Read back ─────────────────────────────────────────────────
        print("\n  TEST 3: read() — read back Form Data sheet")
        df_back = storage.read()
        print(f"  Rows read back: {len(df_back)}")
        assert len(df_back) == len(df_valid)

        # ── TEST 4: Info ──────────────────────────────────────────────────────
        print("\n  TEST 4: info()")
        info = storage.info()
        for k, v in info.items():
            print(f"  {k:<12}: {v}")

        # ── TEST 5: Convenience function ──────────────────────────────────────
        print("\n  TEST 5: save_excel() one-liner")
        path2 = Path(tmpdir) / "quick_save.xlsx"
        r5    = save_excel(df_valid, path2, extra_sheets={"Invalid": df_invalid})
        print(f"  Result : {r5}")
        assert r5.success is True

    print("\n✅ All Excel storage tests passed!")