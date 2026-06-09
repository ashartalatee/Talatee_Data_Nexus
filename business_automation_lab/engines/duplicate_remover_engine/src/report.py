"""
report.py — Report Generator (Excel + Log)
Business Automation Lab | Engine 01: Duplicate Remover

Tugas: Membuat laporan Excel berisi detail duplikat yang ditemukan,
       file yang dihapus, dan ringkasan eksekusi lengkap.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List

from duplicate_detector import DuplicateGroup
from remover import RemovalResult
from scanner import FileRecord

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Buat laporan Excel + log dari hasil engine.

    Contoh penggunaan:
        reporter = ReportGenerator(config)
        reporter.generate(groups, result, all_records)
    """

    def __init__(self, config: dict):
        report_cfg = config.get("report", {})
        paths_cfg = config.get("paths", {})

        self.generate_excel: bool = report_cfg.get("generate_excel", True)
        self.generate_log: bool = report_cfg.get("generate_log", True)
        self.show_space_saved: bool = report_cfg.get("show_space_saved", True)
        self.timestamp_filename: bool = report_cfg.get("timestamp_filename", True)
        self.output_folder = Path(paths_cfg.get("output_folder", "output"))
        self.log_folder = Path(paths_cfg.get("log_folder", "logs"))
        self.dry_run: bool = config.get("action", {}).get("dry_run", True)
        self._run_ts = datetime.now()

    # ------------------------------------------------------------------
    # Public
    # ------------------------------------------------------------------

    def generate(
        self,
        groups: List[DuplicateGroup],
        result: RemovalResult,
        all_records: List[FileRecord],
    ):
        """Generate semua laporan sesuai config."""
        self.output_folder.mkdir(parents=True, exist_ok=True)
        self.log_folder.mkdir(parents=True, exist_ok=True)

        if self.generate_excel:
            self._write_excel(groups, result, all_records)

        if self.generate_log:
            self._write_log(groups, result)

        self._print_console_summary(groups, result, all_records)

    # ------------------------------------------------------------------
    # Excel report
    # ------------------------------------------------------------------

    def _write_excel(
        self,
        groups: List[DuplicateGroup],
        result: RemovalResult,
        all_records: List[FileRecord],
    ):
        """Buat file Excel dengan 3 sheet: Summary, Detail Duplikat, Semua File."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error(
                "openpyxl belum terinstall. Jalankan: pip install openpyxl"
            )
            return

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"
        self._fill_summary_sheet(ws_sum, groups, result, all_records)

        # ── Sheet 2: Duplicate Detail ─────────────────────────────────
        ws_dup = wb.create_sheet("Duplicate Detail")
        self._fill_duplicate_sheet(ws_dup, groups)

        # ── Sheet 3: All Files Scanned ────────────────────────────────
        ws_all = wb.create_sheet("All Files")
        self._fill_all_files_sheet(ws_all, all_records)

        # Simpan file
        filename = self._make_filename("duplicate_report", ".xlsx")
        filepath = self.output_folder / filename
        wb.save(filepath)
        logger.info(f"Laporan Excel disimpan: {filepath}")
        print(f"  📊 Laporan Excel: {filepath}")

    def _fill_summary_sheet(self, ws, groups, result, all_records):
        """Sheet ringkasan eksekusi."""
        from openpyxl.styles import Font, PatternFill, Alignment

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

        # Header
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        ws["A1"] = "DUPLICATE REMOVER ENGINE — LAPORAN EKSEKUSI"
        ws["A1"].font = Font(bold=True, size=14, color="1F3864")
        ws.merge_cells("A1:B1")

        ws["A2"] = f"Dijalankan: {self._run_ts:%Y-%m-%d %H:%M:%S}"
        ws["A2"].font = Font(color="595959", italic=True)
        ws.merge_cells("A2:B2")

        rows = [
            ("HASIL SCAN", ""),
            ("Total file discan", len(all_records)),
            ("Total grup duplikat ditemukan", len(groups)),
            ("Total file duplikat", sum(len(g.duplicates) for g in groups)),
            ("Storage bisa dihemat (MB)", f"{sum(g.wasted_mb for g in groups):.2f} MB"),
            ("", ""),
            ("HASIL EKSEKUSI", ""),
            ("Mode", "DRY RUN (preview)" if self.dry_run else "EKSEKUSI NYATA"),
            ("File diproses/diarsipkan", result.total_removed),
            ("File gagal diproses", result.total_failed),
            ("Storage dibebaskan (MB)", f"{result.mb_freed:.2f} MB"),
        ]

        for i, (label, value) in enumerate(rows, start=4):
            cell_a = ws.cell(row=i, column=1, value=label)
            cell_b = ws.cell(row=i, column=2, value=value)
            if value == "":  # section header
                cell_a.font = Font(bold=True, color="1F3864")
                cell_a.fill = PatternFill("solid", fgColor="DCE6F1")
                cell_b.fill = PatternFill("solid", fgColor="DCE6F1")
            else:
                cell_b.alignment = Alignment(horizontal="right")

    def _fill_duplicate_sheet(self, ws, groups):
        """Sheet detail setiap grup duplikat."""
        from openpyxl.styles import Font, PatternFill, Alignment

        headers = [
            "No", "Grup ID (Hash)", "Status", "Nama File",
            "Path Lengkap", "Ukuran (KB)", "Dimodifikasi"
        ]
        widths = [5, 20, 10, 30, 50, 14, 20]

        # Header row
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = w

        row = 2
        for i, group in enumerate(groups, 1):
            # Master row
            short_id = group.group_id[:16] + "..." if len(group.group_id) > 16 else group.group_id
            m = group.master
            for col, val in enumerate([
                i, short_id, "✅ KEEP", m.name, m.path_str,
                m.size_kb, m.modified_at.strftime("%Y-%m-%d %H:%M")
            ], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor="E2EFDA")  # hijau muda
            row += 1

            # Duplicate rows
            for dupe in group.duplicates:
                for col, val in enumerate([
                    "", short_id, "❌ DUPLIKAT", dupe.name, dupe.path_str,
                    dupe.size_kb, dupe.modified_at.strftime("%Y-%m-%d %H:%M")
                ], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")  # merah muda
                row += 1

            # Baris kosong antar grup
            row += 1

    def _fill_all_files_sheet(self, ws, all_records):
        """Sheet daftar semua file yang discan."""
        from openpyxl.styles import Font, PatternFill, Alignment

        headers = ["No", "Nama File", "Ekstensi", "Ukuran (KB)", "Dimodifikasi", "Status", "Path"]
        widths = [5, 30, 10, 14, 20, 14, 50]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = w

        for i, rec in enumerate(all_records, 2):
            status = "Duplikat" if rec.is_duplicate else "Unik"
            fill_color = "FCE4D6" if rec.is_duplicate else "FFFFFF"
            for col, val in enumerate([
                i - 1, rec.name, rec.extension, rec.size_kb,
                rec.modified_at.strftime("%Y-%m-%d %H:%M"),
                status, rec.path_str
            ], 1):
                cell = ws.cell(row=i, column=col, value=val)
                if rec.is_duplicate:
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill("solid", fgColor=fill_color)

    # ------------------------------------------------------------------
    # Log file
    # ------------------------------------------------------------------

    def _write_log(self, groups: List[DuplicateGroup], result: RemovalResult):
        """Tulis log teks ringkasan eksekusi."""
        filename = self._make_filename("engine_run", ".log")
        filepath = self.log_folder / filename

        lines = [
            "=" * 60,
            "DUPLICATE REMOVER ENGINE — EXECUTION LOG",
            f"Waktu    : {self._run_ts:%Y-%m-%d %H:%M:%S}",
            f"Mode     : {'DRY RUN' if self.dry_run else 'LIVE'}",
            "=" * 60,
            f"Grup duplikat  : {len(groups)}",
            f"File duplikat  : {sum(len(g.duplicates) for g in groups)}",
            f"File diproses  : {result.total_removed}",
            f"File gagal     : {result.total_failed}",
            f"Storage freed  : {result.mb_freed:.2f} MB",
            "",
            "DETAIL DUPLIKAT:",
            "-" * 60,
        ]

        for i, group in enumerate(groups, 1):
            lines.append(f"\n[Grup #{i}] Hash: {group.group_id[:32]}...")
            lines.append(f"  KEEP     : {group.master.path_str}")
            for dupe in group.duplicates:
                lines.append(f"  HAPUS    : {dupe.path_str}")

        if result.failed:
            lines += ["", "FILE GAGAL DIPROSES:", "-" * 60]
            for rec, err in result.failed:
                lines.append(f"  {rec.path_str} — {err}")

        filepath.write_text("\n".join(lines), encoding="utf-8")
        logger.info(f"Log disimpan: {filepath}")
        print(f"  📄 Log file  : {filepath}")

    # ------------------------------------------------------------------
    # Console summary
    # ------------------------------------------------------------------

    def _print_console_summary(self, groups, result, all_records):
        """Print ringkasan ke terminal."""
        total_dupes = sum(len(g.duplicates) for g in groups)
        wasted_mb = sum(g.wasted_mb for g in groups)

        print("\n" + "=" * 60)
        print("RINGKASAN EKSEKUSI")
        print("=" * 60)
        print(f"  File discan        : {len(all_records)}")
        print(f"  Grup duplikat      : {len(groups)}")
        print(f"  File duplikat      : {total_dupes}")
        print(f"  Storage bisa hemat : {wasted_mb:.2f} MB")
        print(f"  File diproses      : {result.total_removed}")
        print(f"  Gagal              : {result.total_failed}")
        if not self.dry_run:
            print(f"  Storage dibebaskan : {result.mb_freed:.2f} MB")
        if self.dry_run:
            print("\n  ⚠️  DRY RUN — tidak ada file yang diubah")
        print("=" * 60 + "\n")

    # ------------------------------------------------------------------
    # Utilities
    # ------------------------------------------------------------------

    def _make_filename(self, prefix: str, ext: str) -> str:
        if self.timestamp_filename:
            ts = self._run_ts.strftime("%Y%m%d_%H%M%S")
            return f"{prefix}_{ts}{ext}"
        return f"{prefix}{ext}"